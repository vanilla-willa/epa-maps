import requests
import json
import time
import openpyxl
from openpyxl import Workbook
import config
import copy
import pandas as pd
from pandas import ExcelWriter
import os.path


class GooglePlaces(object):
    def __init__(self, apiKey, coords, radius, fields):
        super(GooglePlaces, self).__init__()
        self.apiKey = apiKey
        self.coords = coords
        self.radius = radius
        self.fields = fields
        self.columns = ['name', 'types', 'coordinates','address', 'phone', 'website']
 
    # testing function - only returns one result from API
    def get_place(self, input_text):
        endpoint_url = "https://maps.googleapis.com/maps/api/place/findplacefromtext/json"
        params = {
            'key': self.apiKey,
            'input': input_text, 
            'inputtype': 'textquery',
            'fields': ",".join(self.fields),
            'locationbias': "".join( ('circle:', self.radius, '@', self.coords) )
        }
        res = requests.get(endpoint_url, params = params)
        results =  json.loads(res.content)
        return results

    def get_places(self, keyword, **kwargs):
        endpoint_url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
        places = []
        params = {
            'key': self.apiKey,
            'location': self.coords,
            'radius': self.radius,
            'keyword': keyword, 
        }
        if kwargs['category'] is not None:
            print("Extra parameter category found")
            params['type'] = kwargs['category']

        res = requests.get(endpoint_url, params = params)
        results =  json.loads(res.content)
        # print("in json results: ", results)
        places.extend(results['results'])
        time.sleep(2)
        while "next_page_token" in results:
            params['pagetoken'] = results['next_page_token'],
            res = requests.get(endpoint_url, params = params)
            results = json.loads(res.content)
            places.extend(results['results'])
            time.sleep(2)

        return places

    def get_place_details(self, place_id, fields):
        endpoint_url = "https://maps.googleapis.com/maps/api/place/details/json"
        params = {
            'place_id': place_id,
            'fields': ",".join(fields),
            'key': self.apiKey
        }
        res = requests.get(endpoint_url, params = params)
        place_details =  json.loads(res.content)
        return place_details
    
    def process_data(self, keyword, category=None):
        print("keyword being processed: ", keyword)
        print("category being processed: ", category)
        places = self.get_places(keyword, category=category)

        print("====================")
        print(str( len(places) ) + " results received")

        r = 1
        data = []
        place_data = {}
        for place in places:
            r += 1
            details = self.get_place_details(place['place_id'], self.fields)

            name = details['result'].get('name', '')
            place_data['name'] = name

            types = details['result'].get('types', '')
            place_data['types'] = "\n".join(types)

            long_lat = details['result'].get('geometry', '')
            if isinstance(long_lat, dict):
                lat = long_lat['location']['lat']
                lng = long_lat['location']['lng']
                place_data['coordinates'] = "".join( (str(lat), ", ", str(lng)) )

            address = details['result'].get('formatted_address', '')
            place_data['address'] = address

            phone = details['result'].get('formatted_phone_number', '')
            place_data['phone'] = phone

            website = details['result'].get('website', '')
            place_data['website'] = website

            data.append(copy.deepcopy(place_data))
            place_data.clear()

        return data

    def loop_keywords(self, **kwargs):
        dataframes = {}
        if 'category_types' in kwargs:
            for category_type in kwargs['category_types']:
                print("Retrieving results for keyword: ", category_type)
                if '_' in category_type:
                    keyword = category_type.replace('_', '')
                else:
                    keyword = category_type
                
                # data is a list of dictionaries
                data = self.process_data(keyword + ' in East Palo Alto', category_type)
                dataframes[category_type] = pd.DataFrame(data).reindex(columns=self.columns)
                
                print("====== updated dataframe ===========")
                # print(dataframes)

        if 'keywords' in kwargs:
            for keyword in kwargs['keywords']:
                print("Retrieving results for keyword: ", keyword)
                
                # data is a list of dictionaries
                data = self.process_data(keyword + ' in East Palo Alto')
                dataframes[keyword] = pd.DataFrame(data).reindex(columns=self.columns)
                
                print("====== updated dataframe ===========")
                # print(dataframes)

        return dataframes

    def output_to_excel(self, dataframes):
        print("ouputting to excel sheet")
        dest_filename = 'outreach.xlsx'
        engine = 'openpyxl'
        if os.path.isfile(dest_filename):
            wb = openpyxl.load_workbook(dest_filename)
            writer = pd.ExcelWriter(dest_filename, engine=engine, mode='a')
        else:
            wb = Workbook()
            wb.save(dest_filename)
            writer = pd.ExcelWriter(dest_filename, engine=engine)

        writer.book = wb
        writer.sheets = dict( (ws.title, ws) for ws in wb.worksheets)

        for key, df in dataframes.items():
            sheet_name = key
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        print("saving excel sheet")
        wb.save(dest_filename)

class Data(object):
    def __init__(self):
        super(Data, self).__init__()
        self.data = pd.read_excel('outreach.xlsx', sheet_name=None)

    def merge_data(self):
        df = pd.concat(self.data, ignore_index=True)
        consolidated = df.drop_duplicates(subset=["name"], keep='first')
        print(consolidated)

        print("ouputting to excel sheet")
        dest_filename = 'outreach.xlsx'
        engine = 'openpyxl'
        wb = openpyxl.load_workbook(dest_filename)
        writer = pd.ExcelWriter(dest_filename, engine=engine, mode='a')

        writer.book = wb
        writer.sheets = dict( (ws.title, ws) for ws in wb.worksheets)
        consolidated.to_excel(writer, sheet_name='merged', index=False, header=True)
        wb.save(dest_filename)


if __name__ == '__main__':

    coords = '37.475800, -122.131355'
    radius = '3000'
    fields = ['name', 'type', 'geometry/location','formatted_address', 'formatted_phone_number', 'website']
    api = GooglePlaces(config.api_key, coords, radius, fields)

    category_types = ['church', 'convenience_store', 'daycare', 'doctor', 'hospital', 'school', 'supermarket']
    keywords = ['clinic', 'daycare', 'youth organization', 'non-profit organization', 'social services organization', 'grocery store']

    # to do manually: police station, park, library, city hall, government office, gas station, bus station, cafe, book store

    # make each keyword search into a dataframe
    dataframes = api.loop_keywords(keywords=keywords, category_types=category_types)
    api.output_to_excel(dataframes)

    df = Data()
    consolidated = df.merge_data()